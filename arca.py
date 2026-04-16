"""
MIRRA — Generador Libro IVA Digital (ARCA/AFIP)
Acepta: CSV de comprobantes emitidos (ventas) | CSV de comprobantes recibidos (compras)
        Plantilla XLS/XLSX de Contabilium (ventas) | Planilla Tango XLS (compras)
Genera: 4 archivos TXT en formato ARCA — posiciones fijas, encoding Latin-1

Historial de fixes:
  v1 - Versión inicial
  v2 - cod_op ventas=' ', compras='0' (confirmado con TXT reales)
  v3 - leer_csv_afip: soporta 3 variantes (CUIT+headers, solo CUIT, headers directos)
  v4 - Fecha: soporta todos los formatos (DD/MM/YYYY, YYYY-MM-DD, 20260101, etc.)
  v5 - Compras: leer Emisor (proveedor) en lugar de Receptor
  v6 - Compras: cant_alic=0 para facturas sin IVA (exentas)
  v7 - Ventas: cod_doc=96 con DNI 99999999/0 → normalizar a cod_doc=99
  v8 - Ventas: facturas exentas → alícuota 0003 (cant_alic siempre >= 1)
  v9 - Encoding: Latin-1 + transliteración (ñ, tildes) — ARCA rechaza UTF-8
  v10- CUIT: validar dígito verificador, advertir si inválido
"""

import streamlit as st
import pandas as pd
import csv
import os, io, tempfile, zipfile, unicodedata
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

# ─── CONFIG ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MIRRA · Libro IVA Digital",
    page_icon="📒",
    layout="centered",
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #f4f6f8; }
[data-testid="stHeader"]           { background: transparent; }
[data-testid="stSidebar"]          { background: #003d1f; }
[data-testid="stSidebar"] *        { color: #c8e6c9 !important; }
[data-testid="stSidebar"] hr       { border-color: #1b5e20; }
div[data-testid="stButton"] > button {
    background: #2e7d32 !important; color: white !important;
    border: none !important; border-radius: 7px !important;
    font-weight: 600 !important; width: 100%;
    padding: 0.65rem 1.5rem !important;
}
div[data-testid="stButton"] > button:hover { background: #1b5e20 !important; }
div[data-testid="stButton"] > button:disabled { opacity: 0.4; }
div[data-testid="stDownloadButton"] > button {
    background: white !important; color: #2e7d32 !important;
    border: 1.5px solid #2e7d32 !important; border-radius: 7px !important;
    font-weight: 500 !important; width: 100%;
}
div[data-testid="stDownloadButton"] > button:hover { background: #e8f5e9 !important; }
[data-testid="stFileUploader"] {
    border: 2px dashed #81c784; border-radius: 10px;
    padding: 0.5rem; background: #f9fff9;
}
[data-testid="stMetric"] {
    background: white; border: 1px solid #e0e0e0;
    border-radius: 10px; padding: 0.8rem 1rem;
}
details { border: 1px solid #dce8dc !important; border-radius: 9px !important; }
</style>
""", unsafe_allow_html=True)

# ─── TABLAS ARCA ─────────────────────────────────────────────────────────────

TIPOS_CBTE = {
    '1':'001','2':'002','3':'003','4':'004','5':'005',
    '6':'006','7':'007','8':'008','9':'009','10':'010',
    '11':'011','12':'012','13':'013','15':'015','19':'019',
    '20':'020','21':'021','51':'051','52':'052','53':'053',
    '54':'054','60':'060','61':'061','63':'063','64':'064',
}

ALICUOTAS = {
    '0':'0003', '2.5':'0009', '5':'0008',
    '10.5':'0004', '21':'0005', '27':'0006',
}

ALICUOTA_COLS = [
    ('IVA 2,5%',  'Imp. Neto Gravado IVA 2,5%',  '2.5'),
    ('IVA 5%',    'Imp. Neto Gravado IVA 5%',    '5'),
    ('IVA 10,5%', 'Imp. Neto Gravado IVA 10,5%', '10.5'),
    ('IVA 21%',   'Imp. Neto Gravado IVA 21%',   '21'),
    ('IVA 27%',   'Imp. Neto Gravado IVA 27%',   '27'),
]

MONEDAS    = {'$':'PES', 'U$S':'DOL', 'PES':'PES', 'DOL':'DOL'}
CODIGOS_NC_VENTAS  = {'3','8','13','21','53'}
CODIGOS_NC_COMPRAS = {'3','8','13','21','53'}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def to_decimal(s):
    if not s or str(s).strip() in ('','nan','NaN','None'): return Decimal('0')
    s = str(s).strip().replace('.','').replace(',','.')
    try:    return Decimal(s)
    except: return Decimal('0')

def fmt_importe(valor):
    d = Decimal(str(valor)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return str(abs(d)).replace('.','').zfill(15)

def fmt_num(valor, largo):
    s = ''.join(c for c in str(valor or '') if c.isdigit())
    return (s or '0').zfill(largo)

def fmt_alfa(valor, largo):
    s = str(valor) if valor and str(valor) not in ('nan','NaN','None') else ''
    return s[:largo].ljust(largo)

def parse_fecha(s):
    s = str(s).strip()
    if len(s) == 8 and s.isdigit(): return s
    s = s[:10]
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%Y/%m/%d','%d/%m/%y','%d-%m-%y'):
        try:    return datetime.strptime(s, fmt).strftime('%Y%m%d')
        except: pass
    return '00000000'

def tipo_cbte_txt(t):
    return TIPOS_CBTE.get(str(t).strip().split('.')[0], str(t).zfill(3))

def cod_op_ventas(t):
    return 'R' if str(t).strip().split('.')[0] in CODIGOS_NC_VENTAS else ' '

def cod_op_compras(t):
    return 'A' if str(t).strip().split('.')[0] in CODIGOS_NC_COMPRAS else ' '

def moneda_txt(m):
    return MONEDAS.get(str(m).strip(), 'PES')

def tipo_cambio_txt(tc):
    d = to_decimal(tc) or Decimal('1')
    e   = int(d)
    dec = str((d - e).quantize(Decimal('0.000001')))[2:]
    return str(e).zfill(4) + dec.ljust(6,'0')[:6]

def nro_doc(n, largo=20):
    return ''.join(c for c in str(n or '') if c.isdigit()).zfill(largo)

def verificar_cuit(cuit_str):
    """Verifica el dígito verificador del CUIT. Retorna True si es válido."""
    digits = ''.join(c for c in str(cuit_str or '') if c.isdigit())
    if len(digits) != 11:
        return False
    pesos = [5,4,3,2,7,6,5,4,3,2]
    suma  = sum(int(digits[i]) * pesos[i] for i in range(10))
    resto = suma % 11
    dv_calc = 0 if resto == 0 else (9 if resto == 1 else 11 - resto)
    return dv_calc == int(digits[10])

def transliterar(s):
    """Convierte caracteres no-ASCII a su equivalente ASCII (ñ→n, tildes→sin tilde)."""
    nfd = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')

def detectar_periodo(rows):
    for row in rows:
        f = parse_fecha(row.get('Fecha',''))
        if f != '00000000': return f[:6]
    return datetime.now().strftime('%Y%m')

def _alics(row):
    """Extrae lista (cod_alicuota, neto, iva) de una fila."""
    alics = []
    for col_iva, col_neto, alic in ALICUOTA_COLS:
        iva_v  = to_decimal(row.get(col_iva,  0))
        neto_v = to_decimal(row.get(col_neto, 0))
        if neto_v != 0 or iva_v != 0:
            alics.append((alic, neto_v, iva_v))
    if not alics:
        neto_t = to_decimal(row.get('Imp. Neto Gravado Total', 0))
        iva_t  = to_decimal(row.get('Total IVA', 0))
        if neto_t != 0:
            alics = [('21', neto_t, iva_t)]
    return alics

# ─── GENERADORES ─────────────────────────────────────────────────────────────

def generar_ventas(rows):
    """VENTAS_CBTE (266 chars) + VENTAS_ALICUOTAS (62 chars)."""
    cbte_lines, alic_lines, errores = [], [], []

    for i, row in enumerate(rows):
        try:
            fecha     = parse_fecha(row.get('Fecha',''))
            tipo      = str(row.get('Tipo de Comprobante','')).strip().split('.')[0]
            pto_vta   = fmt_num(row.get('Punto de Venta',  0), 5)
            nro_desde = fmt_num(row.get('Número Desde',    0), 20)
            nro_hasta = fmt_num(row.get('Número Hasta', nro_desde), 20)
            moneda    = moneda_txt(row.get('Moneda','$'))
            tc        = tipo_cambio_txt(row.get('Tipo Cambio','1'))

            # Código de documento y número de identificación
            cod_doc   = fmt_num(row.get('Tipo Doc. Receptor', 99), 2)
            _nro_raw  = str(row.get('Nro. Doc. Receptor', '0') or '0').strip()
            _nom_raw  = str(row.get('Denominación Receptor', '') or '').strip()

            _nro_digits  = _nro_raw.lstrip('0')
            _dni_invalido = _nro_raw in ('','0','99999999') or len(_nro_digits) < 7

            if cod_doc == '99' or (cod_doc == '96' and _dni_invalido):
                # Consumidor final sin identificación
                cod_doc = '99'
                nro_id  = '0' * 20
                nombre  = fmt_alfa('VENTAS DEL DIA', 30)
            elif cod_doc == '80' and not verificar_cuit(_nro_raw):
                # CUIT con dígito verificador inválido
                nro_id = '0' * 20
                nombre = fmt_alfa(_nom_raw, 30)
                errores.append(f'Fila {i+1}: CUIT {_nro_raw} inválido ({_nom_raw}) — se puso ceros')
            else:
                nro_id = nro_doc(_nro_raw)
                nombre = fmt_alfa(_nom_raw if _nom_raw else 'VENTAS DEL DIA', 30)

            # Importes
            total      = fmt_importe(to_decimal(row.get('Imp. Total',           0)))
            no_gravado = fmt_importe(to_decimal(row.get('Imp. Neto No Gravado', 0)))
            exentas    = fmt_importe(to_decimal(row.get('Imp. Op. Exentas',     0)))
            otros_trib = fmt_importe(to_decimal(row.get('Otros Tributos',       0)))
            cero15     = fmt_importe(0)
            cod_op     = cod_op_ventas(tipo)

            # Alícuotas
            alics = _alics(row)

            # ARCA exige cant_alic >= 1 en ventas
            # Si no hay IVA pero hay exentas/no gravadas → alícuota 0003 (0%)
            if not alics:
                exentas_v = to_decimal(row.get('Imp. Op. Exentas', 0))
                no_grav_v = to_decimal(row.get('Imp. Neto No Gravado', 0))
                if exentas_v != 0:
                    alics = [('0', exentas_v, Decimal('0'))]
                elif no_grav_v != 0:
                    alics = [('0', no_grav_v, Decimal('0'))]

            cant_alic = str(len(alics)) if alics else '1'

            # CBTE: 266 chars exactos
            linea = (
                fecha + tipo_cbte_txt(tipo) + pto_vta + nro_desde + nro_hasta +
                cod_doc + nro_id + nombre +
                total + no_gravado + cero15 + exentas +
                cero15 + cero15 + cero15 + cero15 +
                moneda + tc + cant_alic + cod_op +
                otros_trib + '00000000'
            )
            assert len(linea) == 266, f"CBTE len={len(linea)}"
            cbte_lines.append(linea)

            # ALÍCUOTAS: 62 chars exactos
            for alic, neto_v, iva_v in alics:
                la = (
                    tipo_cbte_txt(tipo) + pto_vta + nro_desde +
                    fmt_importe(neto_v) + ALICUOTAS.get(alic,'0005') + fmt_importe(iva_v)
                )
                assert len(la) == 62, f"ALIC len={len(la)}"
                alic_lines.append(la)

        except Exception as e:
            errores.append(f'Ventas fila {i+1}: {e}')

    # Ordenar por fecha → tipo → pto_vta → nro (igual que el TXT de referencia)
    orden = sorted(range(len(cbte_lines)), key=lambda i: (
        cbte_lines[i][0:8],   # fecha
        cbte_lines[i][11:16], # pto_vta
        cbte_lines[i][8:11],  # tipo
        cbte_lines[i][16:36], # nro
    ))
    cbte_lines = [cbte_lines[i] for i in orden]

    # Ordenar alícuotas con el mismo orden que los CBTE
    alic_by_cbte = {}
    for la in alic_lines:
        key = (la[0:3], la[3:8], la[8:28])
        alic_by_cbte.setdefault(key, []).append(la)

    alic_lines = []
    for linea in cbte_lines:
        key = (linea[8:11], linea[11:16], linea[16:36])
        for la in alic_by_cbte.get(key, []):
            alic_lines.append(la)

    return cbte_lines, alic_lines, errores


def generar_compras(rows):
    """COMPRAS_CBTE (325 chars) + COMPRAS_ALICUOTAS (84 chars)."""
    cbte_lines, alic_lines, errores = [], [], []

    for i, row in enumerate(rows):
        try:
            fecha    = parse_fecha(row.get('Fecha',''))
            tipo     = str(row.get('Tipo de Comprobante','')).strip().split('.')[0]
            pto_vta  = fmt_num(row.get('Punto de Venta', 0), 5)
            nro_cbte = fmt_num(row.get('Número Desde',   0), 20)
            despacho = fmt_alfa('', 16)
            cod_doc  = fmt_num(row.get('Tipo Doc. Vendedor',
                               row.get('Tipo Doc. Receptor', 80)), 2)
            nro_id   = nro_doc(row.get('Nro. Doc. Vendedor',
                               row.get('Nro. Doc. Receptor', 0)))
            nombre   = fmt_alfa(row.get('Denominación Vendedor',
                                row.get('Denominación Receptor','')), 30)
            moneda   = moneda_txt(row.get('Moneda','$'))
            tc       = tipo_cambio_txt(row.get('Tipo Cambio','1'))

            total      = fmt_importe(to_decimal(row.get('Imp. Total',           0)))
            no_gravado = fmt_importe(to_decimal(row.get('Imp. Neto No Gravado', 0)))
            exentas    = fmt_importe(to_decimal(row.get('Imp. Op. Exentas',     0)))
            cred_fisc  = fmt_importe(to_decimal(row.get('Total IVA',            0)))
            otros_trib = fmt_importe(to_decimal(row.get('Otros Tributos',       0)))
            cero15     = fmt_importe(0)
            cod_op     = cod_op_compras(tipo)

            alics = _alics(row)
            # En compras cant_alic=0 está permitido (facturas exentas)
            cant_alic = str(len(alics)) if alics else '0'

            # CBTE: 325 chars exactos
            linea = (
                fecha + tipo_cbte_txt(tipo) + pto_vta + nro_cbte + despacho +
                cod_doc + nro_id + nombre +
                total + no_gravado + exentas +
                cero15 + cero15 + cero15 + cero15 + cero15 +
                moneda + tc + cant_alic + cod_op +
                cred_fisc + otros_trib +
                fmt_num(0, 11) + fmt_alfa('', 30) + cero15
            )
            assert len(linea) == 325, f"CBTE len={len(linea)}"
            cbte_lines.append(linea)

            # ALÍCUOTAS: 84 chars exactos
            for alic, neto_v, iva_v in alics:
                la = (
                    tipo_cbte_txt(tipo) + pto_vta + nro_cbte +
                    cod_doc + nro_id +
                    fmt_importe(neto_v) + ALICUOTAS.get(alic,'0005') + fmt_importe(iva_v)
                )
                assert len(la) == 84, f"ALIC len={len(la)}"
                alic_lines.append(la)

        except Exception as e:
            errores.append(f'Compras fila {i+1}: {e}')

    # Ordenar por fecha → tipo → pto_vta → nro
    orden = sorted(range(len(cbte_lines)), key=lambda i: (
        cbte_lines[i][0:8],
        cbte_lines[i][8:11],
        cbte_lines[i][11:16],
        cbte_lines[i][16:36],
    ))
    cbte_lines = [cbte_lines[i] for i in orden]

    alic_by_cbte = {}
    for la in alic_lines:
        key = (la[0:3], la[3:8], la[8:28])
        alic_by_cbte.setdefault(key, []).append(la)

    alic_lines = []
    for linea in cbte_lines:
        key = (linea[8:11], linea[11:16], linea[16:36])
        for la in alic_by_cbte.get(key, []):
            alic_lines.append(la)

    return cbte_lines, alic_lines, errores


# ─── LECTORES ─────────────────────────────────────────────────────────────────

def _xls_to_xlsx(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name
    if ext == '.xls':
        out_dir = tempfile.mkdtemp()
        os.system(f'libreoffice --headless --convert-to xlsx "{tmp_path}" --outdir "{out_dir}" 2>/dev/null')
        base = os.path.splitext(os.path.basename(tmp_path))[0]
        tmp_path = os.path.join(out_dir, base + '.xlsx')
    return tmp_path


def leer_xls_ventas(uploaded_file):
    """Plantilla XLS/XLSX de ventas (Contabilium). Hoja 'Comprobantes'."""
    tmp_path = _xls_to_xlsx(uploaded_file)
    df = pd.read_excel(tmp_path, sheet_name='Comprobantes', dtype=str).fillna('')

    rows = []
    for _, r in df.iterrows():
        numero = str(r.get('Número','')).strip()
        pto, nro = numero.split('-',1) if '-' in numero else ('00001', numero)

        cuit_raw = str(r.get('Nro. Documento / Cuit','')).strip()
        try:    cuit_raw = str(int(float(cuit_raw))) if cuit_raw else '0'
        except: cuit_raw = '0'

        rows.append({
            'Fecha':                     str(r.get('Fecha','')).strip()[:10],
            'Tipo de Comprobante':       str(r.get('Tipo comprobante AFIP','')).strip().split('.')[0],
            'Punto de Venta':            pto.lstrip('0') or '0',
            'Número Desde':              nro.lstrip('0') or '0',
            'Número Hasta':              nro.lstrip('0') or '0',
            'Tipo Doc. Receptor':        str(r.get('Tipo documento','99')).strip().split('.')[0],
            'Nro. Doc. Receptor':        cuit_raw,
            'Denominación Receptor':     str(r.get('Razón social','')).strip(),
            'Tipo Cambio':               str(r.get('Cotización','1')).strip() or '1',
            'Moneda':                    '$',
            'IVA 21%':                   str(r.get('IVA 21','')).strip(),
            'Imp. Neto Gravado IVA 21%': str(r.get('Neto Gravado 21','')).strip(),
            'Imp. Neto Gravado Total':   str(r.get('Neto Gravado 21','')).strip(),
            'Imp. Neto No Gravado':      '0',
            'Imp. Op. Exentas':          '0',
            'Otros Tributos':            '0',
            'Total IVA':                 str(r.get('IVA 21','')).strip(),
            'Imp. Total':                str(r.get('Total','')).strip(),
        })
    return rows


def leer_xls_compras(uploaded_file):
    """Plantilla XLS de compras (Planilla Tango). Columna CUIT = proveedor."""
    tmp_path = _xls_to_xlsx(uploaded_file)

    try:
        df = pd.read_excel(tmp_path, sheet_name='Planilla Tango', dtype=str).fillna('')
        fuente = 'Planilla Tango'
    except Exception:
        df = pd.read_excel(tmp_path, sheet_name=0, dtype=str).fillna('')
        fuente = 'hoja 1'

    TIPOS_MAP = {
        'FACTURA A':'1','FACTURA B':'6','FACTURA C':'11',
        'NOTA DEBITO A':'2','NOTA DEBITO B':'7','NOTA DEBITO C':'12',
        'NOTA CREDITO A':'3','NOTA CREDITO B':'8','NOTA CREDITO C':'13',
        'FACTURA':'1',
    }

    rows = []
    for _, r in df.iterrows():
        tipo_txt = str(r.get('TIPO COMPROBANTE','')).strip().upper()
        letra    = str(r.get('LETRA','')).strip().upper()
        clave    = f'{tipo_txt} {letra}'.strip() if letra else tipo_txt
        tipo_afip = TIPOS_MAP.get(clave, TIPOS_MAP.get(tipo_txt, '1'))

        pto = str(r.get('SUCURSAL','1')).strip().split('.')[0]
        nro = str(r.get('COMPROBANTE','')).strip().split('.')[0]

        cuit_raw = str(r.get('CUIT','')).strip()
        try:    cuit_raw = str(int(float(cuit_raw))) if cuit_raw else '0'
        except: cuit_raw = '0'

        iva21_cols  = [0.21,'0.21','IVA 21%','IVA_21','IVA 21']
        neto21_cols = ['NETO GRAV_21','BASE GRAVADA','Neto Gravado 21']
        total_cols  = ['TOTAL FACTURADO','Total','TOTAL']

        iva21  = next((str(r[c]).strip() for c in iva21_cols  if c in r.index and str(r[c]).strip() not in ('','nan')), '0')
        neto21 = next((str(r[c]).strip() for c in neto21_cols if c in r.index and str(r[c]).strip() not in ('','nan')), '0')
        total  = next((str(r[c]).strip() for c in total_cols  if c in r.index and str(r[c]).strip() not in ('','nan')), '0')

        rows.append({
            'Fecha':                     str(r.get('FECHA','')).strip()[:10],
            'Tipo de Comprobante':       tipo_afip,
            'Punto de Venta':            pto.lstrip('0') or '0',
            'Número Desde':              nro.lstrip('0') or '0',
            'Tipo Doc. Vendedor':        '80',
            'Nro. Doc. Vendedor':        cuit_raw,
            'Denominación Vendedor':     str(r.get('NOMBRE PROVEEDOR','')).strip(),
            'Tipo Cambio':               '1',
            'Moneda':                    '$',
            'IVA 21%':                   iva21,
            'Imp. Neto Gravado IVA 21%': neto21,
            'Imp. Neto Gravado Total':   neto21,
            'Imp. Neto No Gravado':      '0',
            'Imp. Op. Exentas':          '0',
            'Otros Tributos':            '0',
            'Total IVA':                 iva21,
            'Imp. Total':                total,
        })
    return rows, fuente


def leer_csv_afip(uploaded_file):
    """
    CSV de AFIP. Maneja 3 variantes:
    A) Col 0 = CUIT + headers en misma fila
    B) Col 0 = solo CUIT, headers en fila siguiente
    C) Headers directamente (sin fila de CUIT) — formato nuevo AFIP
    """
    raw = uploaded_file.read().decode('utf-8-sig')
    sep = ';' if raw.count(';') > raw.count(',') else ','
    all_lines = [l for l in raw.splitlines() if l.strip()]

    def clean(s): return s.strip().strip('"').strip()

    first_cols = [clean(c) for c in all_lines[0].split(sep)]
    col0 = first_cols[0]

    FECHA_COLS = ['Fecha de Emisión','Fecha de emision','Fecha','FECHA','fecha']

    def normalizar(row):
        r = {clean(k): clean(v) for k, v in row.items()}
        if 'Fecha' not in r:
            for fc in FECHA_COLS:
                if fc in r and r[fc]:
                    r['Fecha'] = r[fc]; break
        # CSV recibidos: Emisor = proveedor/vendedor
        if 'Nro. Doc. Emisor' in r:
            r['Tipo Doc. Vendedor']    = r.get('Tipo Doc. Emisor', '80')
            r['Nro. Doc. Vendedor']    = r.get('Nro. Doc. Emisor', '0')
            r['Denominación Vendedor'] = r.get('Denominación Emisor', '')
        return r

    # Variante A: col0 es CUIT + headers
    if col0.isdigit() and len(col0)==11 and len(first_cols)>1 and not first_cols[1].replace(',','').replace('.','').isdigit():
        reader = csv.DictReader(all_lines, delimiter=sep)
        rows = []
        for row in reader:
            r = normalizar(row)
            fecha_key = list(row.keys())[0]
            if 'Fecha' not in r: r['Fecha'] = clean(row[fecha_key])
            rows.append(r)
        return rows

    # Variante B: col0 = solo CUIT
    if col0.isdigit() and len(col0)==11 and len(first_cols)==1:
        reader = csv.DictReader(all_lines[1:], delimiter=sep)
        return [normalizar(row) for row in reader if any(v.strip() for v in row.values())]

    # Variante C: headers directamente
    reader = csv.DictReader(all_lines, delimiter=sep)
    return [normalizar(row) for row in reader if any(v.strip() for v in row.values())]


def leer_archivo_ventas(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext in ('.xls','.xlsx'):
        return leer_xls_ventas(uploaded_file), 'Excel Ventas (Contabilium)'
    return leer_csv_afip(uploaded_file), 'CSV AFIP Ventas'


def leer_archivo_compras(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext in ('.xls','.xlsx'):
        rows, fuente = leer_xls_compras(uploaded_file)
        return rows, f'Excel Compras ({fuente})'
    return leer_csv_afip(uploaded_file), 'CSV AFIP Compras'


def lineas_a_bytes(lines):
    """
    Convierte lista de líneas a bytes.
    ARCA exige ASCII/ISO-8859-1 — NO acepta UTF-8.
    Transliteramos caracteres no-ASCII (ñ→n, tildes→sin tilde).
    """
    lines_ascii = [transliterar(l) for l in lines]
    return ('\r\n'.join(lines_ascii) + '\r\n').encode('latin-1', errors='replace')


def hacer_zip(archivos):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nombre, contenido in archivos.items():
            zf.writestr(nombre, contenido)
    return buf.getvalue()

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 📒 MIRRA")
    st.markdown("**Gestión IVA Digital**")
    st.markdown("---")
    st.markdown("""
**Archivos generados:**
- `VENTAS_CBTE` — 266 chars/línea
- `VENTAS_ALICUOTAS` — 62 chars/línea
- `COMPRAS_CBTE` — 325 chars/línea
- `COMPRAS_ALICUOTAS` — 84 chars/línea
""")
    st.markdown("---")
    st.markdown("""
**Formatos aceptados:**
- CSV del portal AFIP (emitidos/recibidos)
- Plantilla XLS/XLSX (Contabilium)
""")
    st.markdown("---")
    st.caption(f"Estudio Contable MIRRA · {datetime.now().year}")

# ─── HEADER ──────────────────────────────────────────────────────────────────

st.markdown("""
<div style="background:linear-gradient(135deg,#1b5e20 0%,#2e7d32 60%,#388e3c 100%);
            padding:2rem 2rem 1.75rem;border-radius:14px;margin-bottom:2rem;">
    <div style="font-size:0.75rem;font-weight:600;letter-spacing:0.12em;
                color:#a5d6a7;text-transform:uppercase;margin-bottom:0.4rem;">
        Estudio Contable MIRRA
    </div>
    <h1 style="color:white;margin:0;font-size:1.75rem;font-weight:700;letter-spacing:-0.5px;">
        📒 Libro IVA Digital
    </h1>
    <p style="color:#c8e6c9;margin:0.5rem 0 0;font-size:0.9rem;">
        Generador de archivos TXT · Formato ARCA/AFIP · Posiciones fijas
    </p>
</div>
""", unsafe_allow_html=True)

# ─── CARGA DE ARCHIVOS ───────────────────────────────────────────────────────

col1, col2 = st.columns(2)
with col1:
    st.markdown("#### 📤 Ventas")
    st.caption("CSV de **comprobantes emitidos** del portal AFIP")
    file_v = st.file_uploader(
        "ventas", type=['csv','xls','xlsx'],
        key='ventas', label_visibility='collapsed'
    )
with col2:
    st.markdown("#### 📥 Compras *(opcional)*")
    st.caption("CSV de **comprobantes recibidos** del portal AFIP")
    file_c = st.file_uploader(
        "compras", type=['csv','xls','xlsx'],
        key='compras', label_visibility='collapsed'
    )

with st.expander("⚙️ Configuración avanzada"):
    periodo_manual = st.text_input(
        "Forzar período (YYYYMM)",
        placeholder="Ej: 202603 — se detecta automáticamente si lo dejás vacío",
        max_chars=6,
    )

st.divider()

# ─── BOTÓN PROCESAR ──────────────────────────────────────────────────────────

if not file_v:
    st.info("👆 Cargá el CSV de comprobantes emitidos (ventas) para comenzar.")

if st.button("🚀 GENERAR ARCHIVOS TXT", disabled=(file_v is None)):
    archivos_out, todos_errores = {}, []

    with st.spinner("Procesando..."):
        try:
            rows_v, fmt_v = leer_archivo_ventas(file_v)
            periodo = periodo_manual.strip() if periodo_manual.strip() else detectar_periodo(rows_v)

            cbte_v, alic_v, err_v = generar_ventas(rows_v)
            todos_errores.extend(err_v)
            archivos_out[f"LIBRO_IVA_DIGITAL_VENTAS_CBTE_{periodo}.txt"]      = lineas_a_bytes(cbte_v)
            archivos_out[f"LIBRO_IVA_DIGITAL_VENTAS_ALICUOTAS_{periodo}.txt"] = lineas_a_bytes(alic_v)

            fmt_c, n_compras = '', 0
            if file_c:
                rows_c, fmt_c = leer_archivo_compras(file_c)
                n_compras = len(rows_c)
                cbte_c, alic_c, err_c = generar_compras(rows_c)
                todos_errores.extend(err_c)
                archivos_out[f"LIBRO_IVA_DIGITAL_COMPRAS_CBTE_{periodo}.txt"]      = lineas_a_bytes(cbte_c)
                archivos_out[f"LIBRO_IVA_DIGITAL_COMPRAS_ALICUOTAS_{periodo}.txt"] = lineas_a_bytes(alic_c)

            st.session_state.resultado = {
                "archivos":  archivos_out,
                "periodo":   periodo,
                "fmt_v":     fmt_v,
                "fmt_c":     fmt_c,
                "n_ventas":  len(rows_v),
                "n_compras": n_compras,
                "errores":   todos_errores,
            }

        except Exception as e:
            st.error(f"❌ Error crítico: {e}")
            st.stop()

# ─── RESULTADOS ──────────────────────────────────────────────────────────────

if 'resultado' in st.session_state:
    res      = st.session_state.resultado
    periodo  = res["periodo"]
    archivos = res["archivos"]

    if res["errores"]:
        st.warning(f"⚠️ Completado con {len(res['errores'])} advertencia(s).")
        with st.expander("Ver advertencias"):
            for e in res["errores"]:
                st.write(f"- {e}")
    else:
        st.success("✅ Archivos generados correctamente, sin errores.")

    info = f"**Período {periodo[:4]}/{periodo[4:]}** · Ventas: {res['fmt_v']}"
    if res["n_compras"]:
        info += f" · Compras: {res['fmt_c']}"
    st.markdown(info)

    cols_m = st.columns(len(archivos))
    for col, (nombre, contenido) in zip(cols_m, archivos.items()):
        n_reg = contenido.count(b'\r\n')
        label = nombre.replace(f'_{periodo}','').replace('.txt','').replace('LIBRO_IVA_DIGITAL_','')
        col.metric(label, f"{n_reg} reg.")

    st.divider()
    st.markdown("#### 📦 Descargar")

    zip_bytes = hacer_zip(archivos)
    st.download_button(
        label=f"📦 Descargar todo en ZIP  —  LID_{periodo}.zip",
        data=zip_bytes,
        file_name=f"LID_{periodo}.zip",
        mime="application/zip",
    )

    with st.expander("Descargar archivos individuales"):
        cols_d = st.columns(2)
        for i, (nombre, contenido) in enumerate(archivos.items()):
            cols_d[i % 2].download_button(
                label=f"⬇ {nombre}",
                data=contenido,
                file_name=nombre,
                mime="text/plain",
                key=f"dl_{nombre}",
            )

def csv_a_plantilla_tango(csv_bytes: bytes) -> bytes:
    """Convierte CSV de comprobantes emitidos (ARCA) → Plantilla Ventas Tango XLSX."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill('solid', fgColor='B8CCE4')
    HFONT  = Font(name='Calibri', size=11)
    DFONT  = Font(name='Calibri', size=11)
    A_L    = Alignment(horizontal='left',  vertical='bottom')
    A_R    = Alignment(horizontal='right', vertical='bottom')
    TEXT_C  = {1,2,3,6,7,8,9,10,11,20}
    RIGHT_C = {25,26,27,28}
    DATE_C  = {4,5}

    raw   = csv_bytes.decode('utf-8-sig')
    lines = [l for l in raw.splitlines() if l.strip()]
    sep   = ';' if raw.count(';') > raw.count(',') else ','

    def cl(s): return s.strip().strip('"').strip()
    rows_csv = [{cl(k): cl(v) for k,v in r.items()}
                for r in csv.DictReader(lines, delimiter=sep)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comprobantes'
    ws.row_dimensions[1].height = 15.0

    for ci, h in enumerate(_TANGO_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HFONT
        c.alignment = A_R if ci in RIGHT_C else A_L
        if ci in TEXT_C: c.number_format = '@'

    for ci in range(1, len(_TANGO_HEADERS)+1):
        w = _TANGO_COL_WIDTHS.get(get_column_letter(ci))
        if w: ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(rows_csv, 2):
        tipo_raw = r.get('Tipo de Comprobante','').strip()
        lc, td, c3 = _TIPO_TANGO.get(tipo_raw, ('?','FCV',tipo_raw.zfill(3)))

        pto = r.get('Punto de Venta','').strip().zfill(5)
        nro = r.get('Número Desde','').strip().zfill(8)
        fecha = parse_fecha(r.get('Fecha de Emisión',''))

        cod_doc = r.get('Tipo Doc. Receptor','').strip()
        nro_doc = r.get('Nro. Doc. Receptor','').strip()
        nombre  = r.get('Denominación Receptor','').strip()
        try:    cuit_n = int(nro_doc) if nro_doc and nro_doc!='0' else None
        except: cuit_n = None

        cond_iva = 'RI' if cod_doc=='80' else 'CF'
        cae      = _conv_cae(r.get('Cód. Autorización',''))
        cotiz    = float(to_decimal(r.get('Tipo Cambio','1').replace(',','.')) or Decimal('1'))
        neto21   = float(to_decimal(r.get('Imp. Neto Gravado IVA 21%','0')))
        iva21    = float(to_decimal(r.get('IVA 21%','0')))
        total    = float(to_decimal(r.get('Imp. Total','0')))

        vals = [lc, f'{pto}-{nro}', td, fecha, fecha,
                nombre, cod_doc, cuit_n, cond_iva,
                None,'14',None,None,None,None,None,None,None,None,
                c3,'0',None,cae,None,
                cotiz, neto21, iva21, total]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DFONT
            if ci in DATE_C and isinstance(val, datetime):
                c.number_format='m/d/yyyy'; c.alignment=A_L
            elif ci in RIGHT_C: c.alignment=A_R
            elif ci in TEXT_C:  c.number_format='@'; c.alignment=A_L
            else:               c.alignment=A_L

    wc = wb.create_sheet('Configuracion')
    for ci,v in enumerate(['PLANTILLA','COD_MODELO_INGRESO','ID_MODELO_INGRESO','1','2','98'],1):
        wc.cell(row=1,column=ci,value=v)
    for ci,v in enumerate(['Ventas','VENTAS',5,0,1,2],1):
        wc.cell(row=2,column=ci,value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── SECCIÓN: CONVERTIR CSV → PLANTILLA TANGO ────────────────────────────────


# ─── SECCIÓN: CONVERTIR CSV → PLANTILLA TANGO ────────────────────────────────

st.divider()
st.markdown("""
<div style="background:linear-gradient(135deg,#1565C0 0%,#1976D2 100%);
            padding:1.25rem 2rem;border-radius:12px;margin-bottom:1.5rem;">
    <h2 style="color:white;margin:0;font-size:1.3rem;font-weight:600;">
        📋 Convertir CSV ARCA → Plantilla Tango
    </h2>
    <p style="color:#BBDEFB;margin:0.3rem 0 0;font-size:0.85rem;">
        Genera el Excel de importación para Tango a partir del CSV descargado de ARCA
    </p>
</div>
""", unsafe_allow_html=True)

file_csv_tango = st.file_uploader(
    "CSV de comprobantes emitidos (ARCA)",
    type=['csv'],
    key='csv_tango',
    help="Archivo CSV descargado del portal ARCA/AFIP — Comprobantes Emitidos"
)

if st.button("📋 GENERAR PLANTILLA TANGO", disabled=(file_csv_tango is None), key='btn_tango'):
    with st.spinner("Generando plantilla..."):
        try:
            xlsx_bytes = csv_a_plantilla_tango(file_csv_tango.getvalue())
            # Detectar período del CSV
            raw_csv = file_csv_tango.getvalue().decode('utf-8-sig')
            first_line = [l for l in raw_csv.splitlines() if l.strip()][1]
            fecha_str = first_line.split(';')[0].strip().strip('"')
            periodo_tango = parse_fecha(fecha_str)
            periodo_str = periodo_tango.strftime('%Y%m') if periodo_tango else 'YYYYMM'

            st.success(f"✅ Plantilla generada — {len(xlsx_bytes):,} bytes")
            st.download_button(
                label=f"⬇ PlantillaVentas_TANGO_{periodo_str}.xlsx",
                data=xlsx_bytes,
                file_name=f"PlantillaVentas_TANGO_{periodo_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key='dl_tango',
            )
        except Exception as e:
            st.error(f"❌ Error: {e}")
elif file_csv_tango is None:
    st.info("👆 Subí el CSV de comprobantes emitidos para generar la Plantilla Tango.")


# ─── GENERADOR PLANTILLA TANGO ───────────────────────────────────────────────

_TIPO_TANGO = {
    '1':('A','FCV','001'), '2':('A','NDA','002'), '3':('A','NCA','003'),
    '4':('A','RCA','004'), '6':('B','FCV','006'), '7':('B','NDB','007'),
    '8':('B','NCB','008'), '9':('B','RCB','009'), '11':('C','FCV','011'),
    '12':('C','NDC','012'), '13':('C','NCC','013'),
}

_TANGO_HEADERS = [
    'Letra','Número','Tipo de comprobante','Fecha','Fecha contable',
    'Razón social','Tipo documento','Nro. Documento / Cuit','Condición de IVA',
    'Nro. IIBB','Provincia','Sujeto vinculado','Operación habitual','Calle',
    'Localidad','Piso','Departamento','Código postal','Operación sujeto vinculado',
    'Tipo comprobante AFIP','Operación AFIP','Comprobante electrónico','CAE / CAI',
    'Fecha vencimiento CAE / CAI','Cotización','Neto Gravado 21','IVA 21','Total',
]

_TANGO_COL_WIDTHS = {
    'A':5.42,'B':8.29,'C':20.0,'D':6.14,'E':14.29,'F':11.71,'G':15.42,
    'H':21.0,'I':16.29,'J':8.71,'K':9.14,'L':15.71,'M':18.0,'N':5.42,
    'O':9.29,'P':4.71,'Q':13.86,'R':13.0,'S':25.42,'T':21.85,'U':14.57,
    'V':23.86,'W':9.14,'X':26.71,'Y':10.14,'Z':11.43,'AA':11.43,'AB':11.43,
}


def _conv_cae(s):
    s = str(s).strip().replace(',', '.')
    try:    return int(float(s))
    except: return s if s else None


def csv_a_plantilla_tango(csv_bytes: bytes) -> bytes:
    """Convierte CSV de comprobantes emitidos (ARCA) → Plantilla Ventas Tango XLSX."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill('solid', fgColor='B8CCE4')
    HFONT  = Font(name='Calibri', size=11)
    DFONT  = Font(name='Calibri', size=11)
    A_L    = Alignment(horizontal='left',  vertical='bottom')
    A_R    = Alignment(horizontal='right', vertical='bottom')
    TEXT_C  = {1,2,3,6,7,8,9,10,11,20}
    RIGHT_C = {25,26,27,28}
    DATE_C  = {4,5}

    raw   = csv_bytes.decode('utf-8-sig')
    lines = [l for l in raw.splitlines() if l.strip()]
    sep   = ';' if raw.count(';') > raw.count(',') else ','

    def cl(s): return s.strip().strip('"').strip()
    rows_csv = [{cl(k): cl(v) for k,v in r.items()}
                for r in csv.DictReader(lines, delimiter=sep)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comprobantes'
    ws.row_dimensions[1].height = 15.0

    for ci, h in enumerate(_TANGO_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HFONT
        c.alignment = A_R if ci in RIGHT_C else A_L
        if ci in TEXT_C: c.number_format = '@'

    for ci in range(1, len(_TANGO_HEADERS)+1):
        w = _TANGO_COL_WIDTHS.get(get_column_letter(ci))
        if w: ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(rows_csv, 2):
        tipo_raw = r.get('Tipo de Comprobante','').strip()
        lc, td, c3 = _TIPO_TANGO.get(tipo_raw, ('?','FCV',tipo_raw.zfill(3)))

        pto = r.get('Punto de Venta','').strip().zfill(5)
        nro = r.get('Número Desde','').strip().zfill(8)
        fecha = parse_fecha(r.get('Fecha de Emisión',''))

        cod_doc = r.get('Tipo Doc. Receptor','').strip()
        nro_doc = r.get('Nro. Doc. Receptor','').strip()
        nombre  = r.get('Denominación Receptor','').strip()
        try:    cuit_n = int(nro_doc) if nro_doc and nro_doc!='0' else None
        except: cuit_n = None

        cond_iva = 'RI' if cod_doc=='80' else 'CF'
        cae      = _conv_cae(r.get('Cód. Autorización',''))
        cotiz    = float(to_decimal(r.get('Tipo Cambio','1').replace(',','.')) or Decimal('1'))
        neto21   = float(to_decimal(r.get('Imp. Neto Gravado IVA 21%','0')))
        iva21    = float(to_decimal(r.get('IVA 21%','0')))
        total    = float(to_decimal(r.get('Imp. Total','0')))

        vals = [lc, f'{pto}-{nro}', td, fecha, fecha,
                nombre, cod_doc, cuit_n, cond_iva,
                None,'14',None,None,None,None,None,None,None,None,
                c3,'0',None,cae,None,
                cotiz, neto21, iva21, total]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DFONT
            if ci in DATE_C and isinstance(val, datetime):
                c.number_format='m/d/yyyy'; c.alignment=A_L
            elif ci in RIGHT_C: c.alignment=A_R
            elif ci in TEXT_C:  c.number_format='@'; c.alignment=A_L
            else:               c.alignment=A_L

    wc = wb.create_sheet('Configuracion')
    for ci,v in enumerate(['PLANTILLA','COD_MODELO_INGRESO','ID_MODELO_INGRESO','1','2','98'],1):
        wc.cell(row=1,column=ci,value=v)
    for ci,v in enumerate(['Ventas','VENTAS',5,0,1,2],1):
        wc.cell(row=2,column=ci,value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── SECCIÓN: CONVERTIR CSV → PLANTILLA TANGO ────────────────────────────────

